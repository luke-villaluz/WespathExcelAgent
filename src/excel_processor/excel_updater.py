import openpyxl
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime

def update_excel_with_transactions(
    excel_path: Path, 
    transactions: Dict[str, Dict[str, List[float]]],
    date: Optional[str] = None
) -> bool:
    """
    Update Excel file with transaction data.
    
    Args:
        excel_path: Path to the Excel file to update
        transactions: Dictionary in format {'P': {'W': [], 'D': []}, 'I': {'W': [], 'D': []}}
        date: Date string (defaults to today if None)
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_path)
        
        # Set default date if not provided
        if date is None:
            date = datetime.now().strftime("%m/%d/%Y")
        
        # Update Series P worksheet
        if 'ESG016 - Wespath Series P' in workbook.sheetnames:
            update_series_worksheet(workbook['ESG016 - Wespath Series P'], transactions['P'], date)
        else:
            print(f"Warning: Series P worksheet not found in {excel_path}")
        
        # Update Series I worksheet
        if 'ESG017 - Wespath Series I' in workbook.sheetnames:
            update_series_worksheet(workbook['ESG017 - Wespath Series I'], transactions['I'], date)
        else:
            print(f"Warning: Series I worksheet not found in {excel_path}")
        
        # Save the workbook
        workbook.save(excel_path)
        print(f"Successfully updated {excel_path}")
        return True
        
    except Exception as e:
        print(f"Error updating Excel file {excel_path}: {e}")
        return False

def update_series_worksheet(worksheet, series_data: Dict[str, List[float]], date: str) -> None:
    """
    Update a specific series worksheet with transaction data.
    
    Args:
        worksheet: The worksheet to update
        series_data: Dictionary with 'W' (withdrawals) and 'D' (deposits) lists
        date: Date string for the new row
    """
    # Find the next empty row (after the last populated row)
    next_row = find_next_empty_row(worksheet)
    
    # Add date in column A
    worksheet[f'A{next_row}'] = date
    
    # Calculate net cashflow (deposits - withdrawals)
    total_deposits = sum(series_data['D'])
    total_withdrawals = sum(series_data['W'])
    net_cashflow = total_deposits - total_withdrawals
    
    # Format the net cashflow value (negative numbers with parentheses)
    if net_cashflow < 0:
        # Format as negative with parentheses
        formatted_value = f"$ ({abs(net_cashflow):,.2f})"
    else:
        # Format as positive
        formatted_value = f"$ {net_cashflow:,.2f}"
    
    # Add net cashflow to column I (Scheduled Cashflow (Net))
    worksheet[f'I{next_row}'] = formatted_value
    
    print(f"Added row {next_row}: Date={date}, Net Cashflow={formatted_value}")
    print(f"  - Deposits: {total_deposits:,.2f}")
    print(f"  - Withdrawals: {total_withdrawals:,.2f}")

def find_next_empty_row(worksheet) -> int:
    """
    Find the next empty row in the worksheet.
    
    Args:
        worksheet: The worksheet to search
    
    Returns:
        int: Row number of the next empty row
    """
    # Start from row 155 (based on the image showing data in row 155)
    # Look for the first row where column A is empty
    for row in range(155, worksheet.max_row + 10):  # Check a few extra rows
        cell_value = worksheet[f'A{row}'].value
        if cell_value is None or str(cell_value).strip() == '':
            return row
    
    # If no empty row found, return the next row after max_row
    return worksheet.max_row + 1
